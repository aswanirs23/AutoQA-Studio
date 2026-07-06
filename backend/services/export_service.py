"""Serialize ``TestCase`` lists to Excel, CSV, JSON, Markdown, or TestRail-style CSV.

``export_test_cases`` returns raw bytes + MIME type + filename suffix for ``Response`` in ``export`` router.
"""

import csv
import io
import json
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font

from backend.models.test_case import TestCase

ExportFormat = Literal["excel", "csv", "json", "markdown", "testrail"]


def export_test_cases(
    cases: list[TestCase], fmt: ExportFormat, project_name: str = ""
) -> tuple[bytes, str, str]:
    """Return (body_bytes, mime_type, filename_suffix)."""
    if fmt == "json":
        payload = {
            "project": project_name,
            "exported_count": len(cases),
            "test_cases": [_tc_to_dict(tc) for tc in cases],
        }
        data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
        return data, "application/json; charset=utf-8", "json"

    if fmt == "markdown":
        lines = [f"# Test cases{f' — {project_name}' if project_name else ''}", ""]
        for tc in cases:
            steps = "\n".join(f"{i + 1}. {s}" for i, s in enumerate(tc.steps))
            lines.extend(
                [
                    f"## {tc.id}: {tc.title}",
                    "",
                    f"- **Feature:** {tc.feature}",
                    f"- **Type:** {tc.type} | **Priority:** {tc.priority}",
                ]
            )
            if tc.source_ref:
                lines.append(f"- **Source:** {tc.source_ref}")
            if tc.preconditions:
                lines.extend(["", "**Preconditions:**", "", tc.preconditions, ""])
            lines.extend(["**Steps:**", "", steps, "", f"**Expected:** {tc.expected_result}", "", "---", ""])
        data = "\n".join(lines).encode("utf-8")
        return data, "text/markdown; charset=utf-8", "md"

    if fmt == "testrail":
        buf = io.StringIO()
        w = csv.writer(buf)
        # Common TestRail import mapping: Title, Section, Template, Type, Priority, Precondition, Steps, Expected Result
        w.writerow(
            [
                "Title",
                "Section",
                "Template",
                "Type",
                "Priority",
                "Precondition",
                "Steps",
                "Expected Result",
            ]
        )
        for tc in cases:
            steps = "\n".join(f"{i+1}. {s}" for i, s in enumerate(tc.steps))
            w.writerow(
                [
                    tc.title,
                    tc.feature,
                    "Test Case (Text)",
                    tc.type,
                    tc.priority.capitalize(),
                    tc.preconditions,
                    steps,
                    tc.expected_result,
                ]
            )
        data = buf.getvalue().encode("utf-8")
        return data, "text/csv; charset=utf-8", "csv"

    rows = _rows(cases)
    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(
            [
                "ID",
                "Feature",
                "Source",
                "Type",
                "Title",
                "Preconditions",
                "Steps",
                "Expected Result",
                "Priority",
            ]
        )
        for r in rows:
            w.writerow(r)
        data = buf.getvalue().encode("utf-8")
        return data, "text/csv; charset=utf-8", "csv"

    wb = Workbook()
    ws = wb.active
    ws.title = "Test cases"
    headers = [
        "ID",
        "Feature",
        "Source",
        "Type",
        "Title",
        "Preconditions",
        "Steps",
        "Expected Result",
        "Priority",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for i, r in enumerate(rows, start=2):
        for col, val in enumerate(r, start=1):
            ws.cell(row=i, column=col, value=val)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"


def _tc_to_dict(tc: TestCase) -> dict:
    return {
        "id": tc.id,
        "feature": tc.feature,
        "type": tc.type,
        "title": tc.title,
        "preconditions": tc.preconditions,
        "steps": tc.steps,
        "expected_result": tc.expected_result,
        "priority": tc.priority,
        "source_ref": tc.source_ref,
        "created_at": tc.created_at.isoformat() if tc.created_at else None,
    }


def _rows(cases: list[TestCase]) -> list[list[str]]:
    out: list[list[str]] = []
    for tc in cases:
        steps = "\n".join(f"{i+1}. {s}" for i, s in enumerate(tc.steps))
        out.append(
            [
                tc.id,
                tc.feature,
                tc.source_ref or "",
                tc.type,
                tc.title,
                tc.preconditions,
                steps,
                tc.expected_result,
                tc.priority,
            ]
        )
    return out
