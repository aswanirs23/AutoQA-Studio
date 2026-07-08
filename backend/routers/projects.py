"""Project CRUD, context JSON, features, test case CRUD, stats, and dashboard data.

All routes are scoped by ``user_id`` from ``get_current_user_id`` so SQLite rows stay
partitioned per user when JWT auth is enabled.
"""

import io
import json
import logging
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile

from backend.db import get_db
from backend.deps import get_current_user_id
from backend.models.requests import (
    BulkDeleteTestCasesBody,
    CreateProjectBody,
    ProjectDetailResponse,
    ProjectStatsResponse,
    ProjectSummaryResponse,
    SaveAuthBody,
    UpdateContextBody,
    UpdateProjectBody,
    UpdateTestCaseBody,
)
from backend.models.test_case import InputRecord, Project, TestCase
from backend.repositories import feature_repo, input_repo, project_repo, testcase_repo
from backend.services.playwright_login import capture_login_session, mask_auth_config

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/projects", tags=["projects"])


@router.post("", response_model=Project)
async def create_project(
    body: CreateProjectBody,
    user_id: str = Depends(get_current_user_id),
) -> Project:
    async with get_db() as db:
        return await project_repo.create_project(db, user_id, body.name, body.description, body.context)


@router.get("", response_model=list[ProjectSummaryResponse])
async def list_projects(user_id: str = Depends(get_current_user_id)) -> list[ProjectSummaryResponse]:
    async with get_db() as db:
        rows = await project_repo.list_projects(db, user_id)
    return [
        ProjectSummaryResponse(
            id=r["id"],
            name=r["name"],
            description=r["description"],
            user_id=user_id,
            feature_count=r["fc"],
            test_case_count=r["tc"],
            updated_at=r["updated_at"],
        )
        for r in rows
    ]


@router.get("/{project_id}", response_model=ProjectDetailResponse)
async def get_project(
    project_id: str,
    user_id: str = Depends(get_current_user_id),
) -> ProjectDetailResponse:
    async with get_db() as db:
        p = await project_repo.get_project(db, user_id, project_id)
        if not p:
            raise HTTPException(status_code=404, detail="Project not found")
        feats = await feature_repo.list_features(db, user_id, project_id)
    p.auth_config = mask_auth_config(p.auth_config)
    return ProjectDetailResponse(project=p, features=feats)


@router.put("/{project_id}", response_model=Project)
async def update_project(
    project_id: str,
    body: UpdateProjectBody,
    user_id: str = Depends(get_current_user_id),
) -> Project:
    async with get_db() as db:
        ok = await project_repo.update_project(
            db,
            user_id,
            project_id,
            name=body.name,
            description=body.description,
            base_url=body.base_url,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Project not found")
        p = await project_repo.get_project(db, user_id, project_id)
    assert p is not None
    p.auth_config = mask_auth_config(p.auth_config)
    return p


@router.put("/{project_id}/context", response_model=Project)
async def update_context(
    project_id: str,
    body: UpdateContextBody,
    user_id: str = Depends(get_current_user_id),
) -> Project:
    async with get_db() as db:
        ok = await project_repo.update_project_context(db, user_id, project_id, body.context)
        if not ok:
            raise HTTPException(status_code=404, detail="Project not found")
        p = await project_repo.get_project(db, user_id, project_id)
    assert p is not None
    p.auth_config = mask_auth_config(p.auth_config)
    return p


@router.put("/{project_id}/auth")
async def save_project_auth(
    project_id: str,
    body: SaveAuthBody,
    user_id: str = Depends(get_current_user_id),
) -> dict:
    async with get_db() as db:
        existing = await project_repo.get_project_auth(db, user_id, project_id)
        if existing is None:
            raise HTTPException(status_code=404, detail="Project not found")
        password = body.password if body.password else existing.get("password", "")
        cfg = {
            "login_url": body.login_url.strip(),
            "username": body.username,
            "password": password,
            "selectors": body.selectors or {},
            "success_check": body.success_check or "",
            "home_path": (body.home_path or existing.get("home_path", "")).strip(),
            "verified_at": existing.get("verified_at", ""),
            "last_error": existing.get("last_error", ""),
        }
        await project_repo.update_project_auth(db, user_id, project_id, cfg)
    return {"auth_config": mask_auth_config(cfg)}


@router.post("/{project_id}/auth/verify")
async def verify_project_auth(
    project_id: str,
    user_id: str = Depends(get_current_user_id),
) -> dict:
    from datetime import datetime, timezone
    async with get_db() as db:
        proj = await project_repo.get_project(db, user_id, project_id)
        if not proj:
            raise HTTPException(status_code=404, detail="Project not found")
        auth = await project_repo.get_project_auth(db, user_id, project_id)
    if not auth or not auth.get("login_url") or not auth.get("password"):
        raise HTTPException(status_code=400, detail="Set login URL, username, and password first.")
    base_url = (proj.base_url or "").strip().rstrip("/")
    res = await capture_login_session(auth, base_url, project_id)
    auth["verified_at"] = datetime.now(timezone.utc).isoformat() if res["ok"] else auth.get("verified_at", "")
    auth["last_error"] = "" if res["ok"] else (res.get("error") or "Login failed")
    async with get_db() as db:
        await project_repo.update_project_auth(db, user_id, project_id, auth)
    return res


@router.post("/{project_id}/generate-description")
async def generate_description(
    project_id: str,
    file: UploadFile = File(...),
    user_id: str = Depends(get_current_user_id),
) -> dict:
    """Upload a document, extract text, send to LLM to produce a project overview."""
    async with get_db() as db:
        p = await project_repo.get_project(db, user_id, project_id)
        if not p:
            raise HTTPException(status_code=404, detail="Project not found")

    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename")

    suffix = Path(file.filename).suffix.lower()
    data = await file.read()
    max_bytes = 15 * 1024 * 1024
    if len(data) > max_bytes:
        raise HTTPException(status_code=413, detail="File too large (max 15 MB)")

    if suffix == ".pdf":
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(data))
        text = "\n\n".join(pg.extract_text() or "" for pg in reader.pages).strip()
    elif suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                line = "\t".join("" if c is None else str(c) for c in row)
                if line.strip():
                    lines.append(line)
        wb.close()
        text = "\n".join(lines).strip()
    elif suffix == ".docx":
        from docx import Document
        doc = Document(io.BytesIO(data))
        text = "\n\n".join(p.text for p in doc.paragraphs if p.text and p.text.strip()).strip()
    elif suffix in (".txt", ".md", ".csv"):
        text = data.decode("utf-8", errors="replace").strip()
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    if not text:
        raise HTTPException(status_code=422, detail="Could not extract any text from file")

    from backend.config import get_effective_settings
    from backend.prompts.templates import OVERVIEW_SYSTEM_PROMPT, build_overview_generation_prompt
    from backend.services.llm_service import _complete_json

    settings = get_effective_settings()
    user_msg = build_overview_generation_prompt(text)
    try:
        raw = await _complete_json(OVERVIEW_SYSTEM_PROMPT, user_msg, settings, None, None)
        parsed = json.loads(raw)
        overview = parsed.get("overview") or raw
    except Exception as exc:
        logger.warning("LLM overview generation failed: %s", exc)
        raise HTTPException(status_code=502, detail=f"LLM failed to generate overview: {exc!s}")

    return {"overview": overview}


@router.delete("/{project_id}")
async def delete_project(project_id: str, user_id: str = Depends(get_current_user_id)) -> dict:
    async with get_db() as db:
        ok = await project_repo.delete_project(db, user_id, project_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Project not found")
    return {"ok": True}


@router.get("/{project_id}/test-cases", response_model=list[TestCase])
async def list_project_test_cases(
    project_id: str,
    feature_id: str | None = Query(None),
    user_id: str = Depends(get_current_user_id),
) -> list[TestCase]:
    async with get_db() as db:
        p = await project_repo.get_project(db, user_id, project_id)
        if not p:
            raise HTTPException(status_code=404, detail="Project not found")
        if feature_id:
            return await testcase_repo.list_test_cases_for_feature(db, project_id, feature_id)
        return await testcase_repo.list_test_cases_for_project(db, project_id)


@router.get("/{project_id}/stats", response_model=ProjectStatsResponse)
async def project_stats(
    project_id: str,
    user_id: str = Depends(get_current_user_id),
) -> ProjectStatsResponse:
    async with get_db() as db:
        p = await project_repo.get_project(db, user_id, project_id)
        if not p:
            raise HTTPException(status_code=404, detail="Project not found")
        agg = await testcase_repo.aggregate_stats(db, project_id)
    return ProjectStatsResponse(
        total=agg["total"],
        by_type=agg["by_type"],
        by_priority=agg["by_priority"],
        by_feature=agg["by_feature"],
    )


@router.get("/{project_id}/input-history", response_model=list[InputRecord])
async def project_input_history(
    project_id: str,
    limit: int = Query(100, ge=1, le=500),
    user_id: str = Depends(get_current_user_id),
) -> list[InputRecord]:
    async with get_db() as db:
        p = await project_repo.get_project(db, user_id, project_id)
        if not p:
            raise HTTPException(status_code=404, detail="Project not found")
        return await input_repo.list_input_history(db, project_id, limit=limit)


@router.patch("/{project_id}/test-cases/{test_case_id}", response_model=TestCase)
async def patch_test_case(
    project_id: str,
    test_case_id: str,
    body: UpdateTestCaseBody,
    user_id: str = Depends(get_current_user_id),
) -> TestCase:
    async with get_db() as db:
        p = await project_repo.get_project(db, user_id, project_id)
        if not p:
            raise HTTPException(status_code=404, detail="Project not found")
        tc = await testcase_repo.update_test_case(
            db,
            project_id,
            test_case_id,
            title=body.title,
            type=body.type,
            preconditions=body.preconditions,
            steps=body.steps,
            expected_result=body.expected_result,
            priority=body.priority,
            source_ref=body.source_ref,
        )
        if not tc:
            raise HTTPException(status_code=404, detail="Test case not found")
        return tc


@router.delete("/{project_id}/test-cases/{test_case_id}")
async def delete_one_test_case(
    project_id: str,
    test_case_id: str,
    user_id: str = Depends(get_current_user_id),
) -> dict:
    async with get_db() as db:
        p = await project_repo.get_project(db, user_id, project_id)
        if not p:
            raise HTTPException(status_code=404, detail="Project not found")
        ok = await testcase_repo.delete_test_case(db, project_id, test_case_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Test case not found")
    return {"ok": True}


@router.post("/{project_id}/test-cases/bulk-delete")
async def bulk_delete_test_cases(
    project_id: str,
    body: BulkDeleteTestCasesBody,
    user_id: str = Depends(get_current_user_id),
) -> dict:
    async with get_db() as db:
        p = await project_repo.get_project(db, user_id, project_id)
        if not p:
            raise HTTPException(status_code=404, detail="Project not found")
        n = await testcase_repo.delete_test_cases_bulk(db, project_id, body.ids)
    return {"deleted": n}
