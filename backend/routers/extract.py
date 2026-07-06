"""Extract plain text from uploaded documents for project context import."""

from __future__ import annotations

import io
from pathlib import Path

from fastapi import APIRouter, File, HTTPException, UploadFile

router = APIRouter(tags=["extract"])

# Guard against huge uploads (context is meant to be human-sized text).
MAX_BYTES = 15 * 1024 * 1024


def _extract_pdf(data: bytes) -> str:
    from PyPDF2 import PdfReader

    reader = PdfReader(io.BytesIO(data))
    parts: list[str] = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            parts.append(t)
    return "\n\n".join(parts).strip()


def _extract_xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for sheet in wb.worksheets:
            lines.append(f"## {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                line = "\t".join("" if c is None else str(c) for c in row)
                if line.strip():
                    lines.append(line)
    finally:
        wb.close()
    return "\n".join(lines).strip()


def _extract_docx(data: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(data))
    paras = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
    return "\n\n".join(paras).strip()


@router.post("/extract-text")
async def extract_text(file: UploadFile = File(...)) -> dict[str, str]:
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename")
    suffix = Path(file.filename).suffix.lower()
    if suffix not in (".pdf", ".xlsx", ".docx"):
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type for server extraction (use .pdf, .xlsx, or .docx)",
        )
    data = await file.read()
    if len(data) > MAX_BYTES:
        raise HTTPException(status_code=413, detail="File too large (max 15 MB)")
    try:
        if suffix == ".pdf":
            text = _extract_pdf(data)
        elif suffix == ".xlsx":
            text = _extract_xlsx(data)
        else:
            text = _extract_docx(data)
    except HTTPException:
        raise
    except Exception as e:  # noqa: BLE001 — return safe message to client
        raise HTTPException(status_code=422, detail=f"Could not extract text: {e!s}") from e
    return {"text": text or ""}
